import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "カブ隊年間スケジュール"

# ---- カラー定義 ----
HEADER_BG   = "FFF2CC"   # 黄色（カブカラー）
MONTH_BG    = "FFE0B2"   # オレンジ
CAMP_BG     = "C8E6C9"   # 緑（キャンプ）
SERVICE_BG  = "BBDEFB"   # 青（奉仕）
EVENT_BG    = "F8BBD0"   # ピンク（行事）
NORMAL_BG   = "FFFFFF"
TITLE_BG    = "FF8F00"   # 濃いオレンジ

thin = Side(style="thin", color="CCCCCC")
thick = Side(style="medium", color="888888")
thin_border  = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

def hfill(color):
    return PatternFill("solid", fgColor=color)

def cell_style(ws, row, col, value, bg=NORMAL_BG, bold=False,
               fontsize=11, align="center", wrap=True, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="メイリオ", size=fontsize, bold=bold, color=color)
    c.fill = hfill(bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                             wrap_text=wrap)
    c.border = thin_border
    return c

# ---- タイトル行 ----
ws.merge_cells("A1:H1")
t = ws.cell(row=1, column=1,
            value="🐯 カブ隊 年間活動スケジュール（2026年度）")
t.font = Font(name="メイリオ", size=16, bold=True, color="FFFFFF")
t.fill = hfill(TITLE_BG)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 36

# ---- ヘッダー行 ----
headers = ["月", "日程（予定）", "活動名", "種別", "場所", "持ち物", "担当", "備考"]
for col, h in enumerate(headers, 1):
    cell_style(ws, 2, col, h, bg=HEADER_BG, bold=True, fontsize=12)
ws.row_dimensions[2].height = 24

# ---- スケジュールデータ ----
# (月, 日程, 活動名, 種別, 場所, 持ち物, 担当, 備考)
schedule = [
    # 4月
    ("4月", "4/5（日）",  "入隊式・対面式",      "行事",   "公民館ホール",   "制服・ネッカチーフ", "隊長",   "新入隊員歓迎"),
    ("",    "4/19（日）", "春のハイキング",       "ハイク", "○○山・里山コース","弁当・水筒・雨具",  "副長",   "約5km"),
    # 5月
    ("5月", "5/3（日）",  "こどもの日デイキャンプ","キャンプ","○○キャンプ場",  "着替え・寝袋不要",  "全員",   "カレー作り"),
    ("",    "5/17（日）", "工作・ペーパークラフト","集会",   "公民館",         "はさみ・のり",      "副長A",  ""),
    # 6月
    ("6月", "6/7（日）",  "奉仕活動（清掃）",    "奉仕",   "○○公園",        "軍手・マスク",      "副長B",  "地域貢献"),
    ("",    "6/21（日）", "救急・応急処置訓練",  "集会",   "公民館",         "包帯・三角巾",      "隊長",   "安全章取得"),
    # 7月
    ("7月", "7/19（日）", "夏キャンプ事前準備",  "集会",   "公民館",         "筆記用具",          "副長A",  "持ち物確認"),
    ("",    "7/24（金）\n〜7/26（日）", "夏季キャンプ", "キャンプ", "○○自然の家", "寝袋・着替え3日分・水筒", "全員", "2泊3日"),
    # 8月
    ("8月", "8/2（日）",  "奉仕活動（ゴミ拾い）","奉仕",   "○○川沿い",      "軍手・ゴミ袋",      "副長B",  "夏休み地域奉仕"),
    ("",    "8/23（日）", "水遊び・野外料理",    "キャンプ","○○川",          "水着・タオル",      "隊長",   "安全確認徹底"),
    # 9月
    ("9月", "9/6（日）",  "秋のハイキング",      "ハイク", "△△峠コース",    "弁当・水筒・雨具",  "副長A",  "約7km・紅葉前"),
    ("",    "9/20（日）", "地図・コンパス訓練",  "集会",   "○○広場",        "コンパス（貸出可）","副長B",  "地図読み基礎"),
    # 10月
    ("10月","10/4（日）", "デイキャンプ・工作",  "キャンプ","○○森林公園",   "軍手・着替え",      "全員",   "木工クラフト"),
    ("",    "10/18（日）","地区カブラリー",       "行事",   "市民運動場",     "制服・弁当",        "隊長",   "他隊との交流"),
    # 11月
    ("11月","11/1（日）", "奉仕活動（落葉清掃）","奉仕",   "○○神社境内",    "軍手・ほうき",      "副長B",  ""),
    ("",    "11/15（日）","ビーバー隊との交流会", "行事",   "公民館",         "制服",              "隊長",   "下の学年と交流"),
    # 12月
    ("12月","12/6（日）", "クリスマス集会・工作","集会",   "公民館",         "材料費200円",       "副長A",  "プレゼント交換"),
    ("",    "12/20（日）","年末奉仕・大掃除",    "奉仕",   "公民館・周辺",   "雑巾・バケツ",      "全員",   "1年の感謝を込めて"),
    # 1月
    ("1月", "1/10（日）", "新年集会・もちつき",  "行事",   "○○公民館",      "制服・エプロン",    "隊長",   "伝統行事"),
    ("",    "1/24（日）", "冬のハイキング",       "ハイク", "○○里山",        "防寒着・弁当",      "副長A",  "防寒対策必須"),
    # 2月
    ("2月", "2/7（日）",  "節分・豆まき集会",    "集会",   "公民館",         "豆（用意します）",  "副長B",  "工作あり"),
    ("",    "2/21（日）", "スキー・雪遊び",       "キャンプ","○○スキー場",   "スキーウェア・手袋","全員",   "希望者参加"),
    # 3月
    ("3月", "3/7（日）",  "進級・卒業準備集会",  "集会",   "公民館",         "制服",              "隊長",   "バッジ確認"),
    ("",    "3/21（日）", "卒業式・お別れキャンプ","行事", "○○キャンプ場",  "制服・弁当",        "全員",   "ボーイ隊へ"),
]

# 種別→背景色マッピング
type_color = {
    "キャンプ": CAMP_BG,
    "奉仕":    SERVICE_BG,
    "行事":    EVENT_BG,
    "ハイク":  "E1BEE7",
    "集会":    NORMAL_BG,
}

row = 3
prev_month = None
for data in schedule:
    month, date, name, kind, place, items, leader, note = data
    bg = type_color.get(kind, NORMAL_BG)

    # 月セルは同じ月なら空白
    month_val = month if month != "" else ""
    cell_style(ws, row, 1, month_val,  bg=MONTH_BG, bold=True, fontsize=12)
    cell_style(ws, row, 2, date,       bg=bg)
    cell_style(ws, row, 3, name,       bg=bg, bold=True, align="left")
    cell_style(ws, row, 4, kind,       bg=bg)
    cell_style(ws, row, 5, place,      bg=bg, align="left")
    cell_style(ws, row, 6, items,      bg=bg, align="left", fontsize=10)
    cell_style(ws, row, 7, leader,     bg=bg)
    cell_style(ws, row, 8, note,       bg=bg, align="left", fontsize=10)
    ws.row_dimensions[row].height = 36
    row += 1

# ---- 列幅調整 ----
col_widths = [8, 18, 24, 8, 20, 28, 10, 22]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# ---- 凡例シート ----
ws2 = wb.create_sheet("凡例・注意事項")
ws2.column_dimensions["A"].width = 16
ws2.column_dimensions["B"].width = 40

legend_title = ws2.cell(row=1, column=1, value="種別カラー凡例")
legend_title.font = Font(name="メイリオ", size=13, bold=True)
ws2.merge_cells("A1:B1")

legends = [
    ("キャンプ",   CAMP_BG,    "キャンプ・デイキャンプ"),
    ("奉仕",       SERVICE_BG, "地域奉仕活動"),
    ("行事",       EVENT_BG,   "式典・イベント"),
    ("ハイク",     "E1BEE7",   "ハイキング"),
    ("集会",       NORMAL_BG,  "通常集会・訓練"),
]
for i, (label, color, desc) in enumerate(legends, 2):
    c1 = ws2.cell(row=i, column=1, value=label)
    c1.fill = hfill(color)
    c1.font = Font(name="メイリオ", size=11, bold=True)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c1.border = thin_border
    c2 = ws2.cell(row=i, column=2, value=desc)
    c2.fill = hfill(color)
    c2.font = Font(name="メイリオ", size=11)
    c2.alignment = Alignment(horizontal="left", vertical="center")
    c2.border = thin_border
    ws2.row_dimensions[i].height = 24

notes_row = len(legends) + 3
ws2.cell(row=notes_row, column=1, value="注意事項").font = Font(
    name="メイリオ", size=12, bold=True)
notes = [
    "・日程は天候・会場都合により変更の場合があります。",
    "・制服着用の活動は「制服・ネッカチーフ着用」が基本です。",
    "・参加費が発生する活動は事前にお知らせします。",
    "・緊急連絡先は各家庭で確認しておいてください。",
    "・活動保険（スポーツ安全保険）への加入が必須です。",
]
for j, n in enumerate(notes, notes_row + 1):
    c = ws2.cell(row=j, column=1, value=n)
    c.font = Font(name="メイリオ", size=10)
    ws2.merge_cells(f"A{j}:B{j}")

out = "/home/user/pubcode/カブ隊年間スケジュール2026.xlsx"
wb.save(out)
print(f"保存完了: {out}")
