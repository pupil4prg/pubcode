import openpyxl
from openpyxl import load_workbook, Workbook

# 入力となるExcelファイルを読み込む
input_file = 'input.xlsx'
wb = load_workbook(input_file)

# 出力用の新しいExcelブックを作成
new_wb = Workbook()
# デフォルトで作成されるシートを削除
new_wb.remove(new_wb.active)

# 各シートを処理
for sheet in wb.worksheets:
    ws = sheet
    min_row = ws.min_row
    max_row = ws.max_row
    min_col = ws.min_column
    max_col = ws.max_column

    num_rows = max_row - min_row + 1
    num_cols = max_col - min_col + 1

    # セルの罫線情報を取得
    cell_borders = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        row_borders = []
        for cell in row:
            borders = {
                'left': cell.border.left.style,
                'right': cell.border.right.style,
                'top': cell.border.top.style,
                'bottom': cell.border.bottom.style
            }
            row_borders.append(borders)
        cell_borders.append(row_borders)

    rectangles = []
    visited = [[False]*num_cols for _ in range(num_rows)]
    for i in range(num_rows):
        for j in range(num_cols):
            if visited[i][j]:
                continue
            borders = cell_borders[i][j]
            if borders['top'] not in [None, 'none'] and borders['left'] not in [None, 'none']:
                # 矩形の探索を開始
                max_height = 1
                max_width = 1
                # 横方向に拡張
                for w in range(1, num_cols - j):
                    if cell_borders[i][j + w]['top'] not in [None, 'none']:
                        max_width += 1
                    else:
                        break
                # 縦方向に拡張
                for h in range(1, num_rows - i):
                    if cell_borders[i + h][j]['left'] not in [None, 'none']:
                        max_height += 1
                    else:
                        break
                # 右と下の罫線を確認
                right_borders_ok = all(cell_borders[i + k][j + max_width - 1]['right'] not in [None, 'none'] for k in range(max_height))
                bottom_borders_ok = all(cell_borders[i + max_height -1][j + k]['bottom'] not in [None, 'none'] for k in range(max_width))
                if right_borders_ok and bottom_borders_ok:
                    rectangle = (i, j, max_height, max_width)
                    rectangles.append(rectangle)
                    # 矩形内のセルを訪問済みにする
                    for x in range(i, i + max_height):
                        for y in range(j, j + max_width):
                            visited[x][y] = True

    # 抽出した矩形領域を新しいシートとして追加
    for idx, rect in enumerate(rectangles):
        start_row, start_col, height, width = rect
        data = []
        for i in range(start_row, start_row + height):
            row_data = []
            for j in range(start_col, start_col + width):
                cell = ws.cell(row=min_row + i, column=min_col + j)
                row_data.append(cell.value)
            data.append(row_data)
        
        # 矩形領域の左上のセルの一つ上のセルを確認
        header = None
        cell_above_row = min_row + start_row - 1
        if cell_above_row >= 1:
            cell_above = ws.cell(row=cell_above_row, column=min_col + start_col)
            if cell_above.value is not None and isinstance(cell_above.value, str):
                header = cell_above.value

        new_sheet_name = f"{ws.title}_Table_{idx + 1}"
        new_ws = new_wb.create_sheet(title=new_sheet_name)

        current_row = 1
        if header:
            new_ws.cell(row=current_row, column=1, value=header)
            current_row += 1  # 次の行から表を開始
        # データを書き込む
        for r_idx, row_data in enumerate(data, current_row):
            for c_idx, value in enumerate(row_data, 1):
                new_ws.cell(row=r_idx, column=c_idx, value=value)

# 新しいブックを保存
output_file = 'output.xlsx'
new_wb.save(output_file)
